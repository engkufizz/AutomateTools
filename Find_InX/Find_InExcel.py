import os
from openpyxl import load_workbook

def find_string_in_xlsx(directory, search_string, output_file, match_whole_word=False):
    # List to store the names of files containing the search string
    files_with_string = []

    # Walk through the directory and its subdirectories
    for root, _, files in os.walk(directory):
        for filename in files:
            # Check if the file is an .xlsx file
            if filename.endswith('.xlsx'):
                file_path = os.path.join(root, filename)
                try:
                    # Load the workbook
                    workbook = load_workbook(file_path, data_only=True)
                    found = False

                    # Iterate over each sheet in the workbook
                    for sheet in workbook.worksheets:
                        # Iterate over each row in the sheet
                        for row in sheet.iter_rows(values_only=True):
                            for cell in row:
                                if cell is not None:
                                    cell_text = str(cell).lower()
                                    search_text = search_string.lower()

                                    if match_whole_word:
                                        # Split the cell text into words and check for exact matches
                                        words = cell_text.split()
                                        if search_text in words:
                                            found = True
                                            break
                                    else:
                                        if search_text in cell_text:
                                            found = True
                                            break
                            if found:
                                break
                        if found:
                            break

                    if found:
                        files_with_string.append(file_path)

                except Exception as e:
                    print(f"Error reading {filename}: {e}")

    # Write the results to the output file
    with open(output_file, 'w') as f:
        for file in files_with_string:
            f.write(file + '\n')

# Define the directory containing the .xlsx files
directory = r'<ENTER_DIRECTORY_PATH_HERE>'
# Define the string to search for
search_string = '<ENTER_SEARCH_STRING_HERE>'
# Define the output file
output_file = 'result.txt'

# Call the function
find_string_in_xlsx(directory, search_string, output_file, match_whole_word=False)

print(f"Search complete. Results saved to {output_file}.")

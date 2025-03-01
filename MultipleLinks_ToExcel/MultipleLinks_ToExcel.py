#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import shutil
from openpyxl import load_workbook

def add_topology_links(excel_file, source_image_folder):
    """
    Add topology links to Excel file using relative paths and create a local image folder.
    
    Args:
        excel_file (str): Path to the Excel file
        source_image_folder (str): Source folder containing the topology images
    """
    try:
        # Load the workbook and select the active worksheet
        wb = load_workbook(excel_file)
        ws = wb.active

        # Get the directory where the Excel file is located
        excel_dir = os.path.dirname(os.path.abspath(excel_file))
        
        # Define the relative image folder name and create full path
        relative_image_folder = "topology_images"
        local_image_folder = os.path.join(excel_dir, relative_image_folder)
        
        # Create the local image folder if it doesn't exist
        os.makedirs(local_image_folder, exist_ok=True)
        
        # Get header from the first row
        header = [cell.value for cell in ws[1]]
        
        # Check if "Topology" header exists; if not, add it in the last column
        if "Topology" in header:
            topology_col = header.index("Topology") + 1
        else:
            topology_col = ws.max_column + 1
            ws.cell(row=1, column=topology_col, value="Topology")
        
        # Counter for tracking processed files
        files_processed = 0
        files_not_found = 0
        
        # Loop over each row starting from row 2 (exclude header)
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
            upe_value = row[0].value  # Assuming UPE value is in column A
            if not upe_value:
                continue
            
            # Build the filename
            image_filename = f"topology_{upe_value}.png"
            source_image_path = os.path.join(source_image_folder, image_filename)
            target_image_path = os.path.join(local_image_folder, image_filename)
            
            # Get the cell for the topology link
            cell = ws.cell(row=row[0].row, column=topology_col)
            
            # Check if source image exists
            if os.path.exists(source_image_path):
                # Copy image to local folder
                shutil.copy2(source_image_path, target_image_path)
                
                # Create relative path link
                relative_link = f"./{relative_image_folder}/{image_filename}"
                
                cell.value = "View Topology"
                cell.hyperlink = relative_link
                cell.style = "Hyperlink"
                files_processed += 1
            else:
                cell.value = "File not found"
                files_not_found += 1
        
        # Save the updated workbook
        file_name, file_ext = os.path.splitext(excel_file)
        updated_file = f"{file_name}_with_topology{file_ext}"
        wb.save(updated_file)
        
        print(f"\nProcess completed successfully!")
        print(f"Files processed: {files_processed}")
        print(f"Files not found: {files_not_found}")
        print(f"Updated Excel file saved as: {updated_file}")
        print(f"Images folder created at: {local_image_folder}")
        
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == '__main__':
    # Define the Excel file and the topology images folder
    excel_file = "upe_pea_analysis_combined_20250227_095134.xlsx"
    source_image_folder = r"d:\XXXXX\Projects\GenerateISIS_TOPO\topology_output"
    add_topology_links(excel_file, source_image_folder)
